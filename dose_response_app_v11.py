import io
import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter, ScalarFormatter, LogLocator, NullFormatter
from scipy.optimize import curve_fit
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage  # Excel画像挿入用

# ==== 初期設定 ====
try:
    st.cache_data.clear()
    st.cache_resource.clear()
except Exception:
    pass

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 10,
    "axes.labelsize": 12,
    "axes.titlesize": 14,
    "legend.fontsize": 10,
    "xtick.labelsize": 12,
    "ytick.labelsize": 12
})

st.set_page_config(page_title="Dose–Response Analysis", layout="wide")
st.title("Dose–Response Curve & EC50/EC90 Calculator [4PL model: Bottom + (Top - Bottom) / (1 + (X/EC50)^Hill)]")

# ==== モデル関数 ====
def four_param_logistic(x, bottom, top, ec50, hill):
    """4PL model: Bottom + (Top - Bottom) / (1 + (X/EC50)^Hill)"""
    return bottom + (top - bottom) / (1 + (x / ec50) ** hill)

def three_param_logistic(x, top, ec50, hill):
    """3PL model: Bottom fixed to 0"""
    bottom = 0.0
    return bottom + (top - bottom) / (1 + (x / ec50) ** hill)

def r_squared(y_obs, y_fit):
    ss_res = np.nansum((y_obs - y_fit) ** 2)
    ss_tot = np.nansum((y_obs - np.nanmean(y_obs)) ** 2)
    return 1 - ss_res / ss_tot if ss_tot > 0 else np.nan

# ==== EC90（下降型4PLの90%効果側） ====
def calc_ec90(ec50, hill):
    """EC90 = EC50 * 9^(1/Hill)"""
    return ec50 * (9 ** (1.0 / hill))

# ==== フィッティング ====
def fit_model(x, y, model="4PL"):
    mask = np.isfinite(x) & np.isfinite(y) & (x > 0)
    x, y = x[mask], y[mask]
    bottom_guess, top_guess = np.min(y), np.max(y)
    half = bottom_guess + 0.5 * (top_guess - bottom_guess)
    idx = int(np.argmin(np.abs(y - half)))
    ec50_guess = x[idx]
    hill_guess = 1.0
    if model == "4PL":
        p0 = [bottom_guess, top_guess, ec50_guess, hill_guess]
        popt, _ = curve_fit(four_param_logistic, x, y, p0=p0, maxfev=20000)
    else:
        p0 = [top_guess, ec50_guess, hill_guess]
        popt, _ = curve_fit(three_param_logistic, x, y, p0=p0, maxfev=20000)
    return popt

# ==== Excel出力（空行区切り＋モデル数式追記付き） ====
def to_excel_multi(df_results=None, df_all=None, df_mean_sd=None, plot_png_buffer=None, model_formula=None) -> bytes:
    tmp = io.BytesIO()
    with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
        # --- Fit Results ---
        if df_results is not None and not df_results.empty:
            df_results.to_excel(writer, index=False, sheet_name="Fit Results")

        # --- Fitted Curve (Dataset区切りあり) ---
        if df_all is not None and not df_all.empty:
            fitted_parts = []
            for ds, df_ds in df_all.groupby("Dataset"):
                fitted_parts.append(df_ds)
                # 空行（列数に合わせた空セル）を追加
                fitted_parts.append(pd.DataFrame([[""] * len(df_ds.columns)], columns=df_ds.columns))
            df_concat = pd.concat(fitted_parts, ignore_index=True)
            df_concat.to_excel(writer, index=False, sheet_name="Fitted Curve")

        # --- Mean_SD + Raw Data (Dataset区切りあり) ---
        if df_mean_sd is not None and not df_mean_sd.empty:
            mean_parts = []
            for ds, df_ds in df_mean_sd.groupby("Dataset"):
                mean_parts.append(df_ds)
                mean_parts.append(pd.DataFrame([[""] * len(df_ds.columns)], columns=df_ds.columns))
            df_concat2 = pd.concat(mean_parts, ignore_index=True)
            df_concat2.to_excel(writer, index=False, sheet_name="Mean_SD + Raw Data")

    # --- モデル式追記 ---
    tmp.seek(0)
    wb = load_workbook(tmp)
    if "Fit Results" in wb.sheetnames:
        ws = wb["Fit Results"]
        row = ws.max_row + 2
        ws.cell(row=row, column=1, value="Model Formula")
        ws.cell(row=row + 1, column=1, value=model_formula)

    # --- 画像埋め込み ---
    if plot_png_buffer is not None:
        ws_plot = wb.create_sheet("Dose-Response Plot")
        plot_png_buffer.seek(0)
        img = XLImage(plot_png_buffer)
        img.width = 640
        img.height = 400
        ws_plot.add_image(img, "B2")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ==== サイドバー ====
st.sidebar.header("Settings")
model = st.sidebar.selectbox("Model", ["4PL", "3PL (Bottom=0)"], index=0)
x_unit = st.sidebar.text_input("Peptide conc.", "nM")
response_label = st.sidebar.text_input("Response label", "RLU")
x_axis_format = st.sidebar.radio("X-axis format", ["normal", "scientific (×10^n)"], index=0)
y_axis_format = st.sidebar.radio("Y-axis format", ["normal", "scientific (×10^n)"], index=1)
manual_scale = st.sidebar.checkbox("🧭 Manually set axis limits", value=False)

# === EC50/R² 表示オプション ===
st.sidebar.markdown("---")
st.sidebar.markdown("### Display options")
show_ec50_line = st.sidebar.checkbox("Show EC50 line", value=False)
show_ec50_list = st.sidebar.checkbox("Show EC50 + R² list (auto adjust)", value=False)

# ==== モデル数式 ====
if model.startswith("4PL"):
    model_formula = "Y = Bottom + (Top - Bottom) / (1 + (X/EC50)^Hill)"
else:
    model_formula = "Y = Top / (1 + (X/EC50)^Hill)"

# ==== データ入力 ====
st.subheader("1) Upload one or more CSV files")
uploads = st.file_uploader("Select CSV files", type=["csv"], accept_multiple_files=True)
example = st.toggle("Use example data", value=(len(uploads) == 0))

fig, ax = plt.subplots(figsize=(8, 5))
colors = plt.cm.tab10.colors
results, all_fitted, mean_sd_list = [], [], []

if example:
    uploads = [("Example", pd.DataFrame({
        "concentration": [1000, 300, 100, 30, 10, 3, 1, 0.3, 0.1, 0.03, 0.01, 0.003, 0.001],
        "Y1": [674558, 633720, 592204, 465966, 453225, 219758, 84598, 19263, 2748, 1200, 800, 600, 500],
        "Y2": [613413, 595589, 639103, 489596, 347633, 204157, 69535, 21848, 1858, 1100, 700, 500, 400],
        "Y3": [606161, 662878, 595337, 498276, 323940, 193936, 59538, 17116, 3317, 1000, 600, 400, 300]
    }))]

for i, f in enumerate(uploads):
    label, data = (f if example else (f.name, pd.read_csv(f)))
    label_clean = label.replace(".csv", "").replace(" data", "")

    x_col = st.selectbox(f"[{label_clean}] Concentration column", data.columns, index=0, key=f"x_{i}")
    y_cols = st.multiselect(f"[{label_clean}] Response columns",
                            [c for c in data.columns if c != x_col],
                            default=[c for c in data.columns if c != x_col],
                            key=f"y_{i}")
    if not y_cols:
        continue

    x = data[x_col].to_numpy(dtype=float)
    y_mean = data[y_cols].mean(axis=1).to_numpy(dtype=float)
    y_sd = data[y_cols].std(axis=1).to_numpy(dtype=float)

    try:
        popt = fit_model(x, y_mean, model=("4PL" if model.startswith("4PL") else "3PL"))
    except Exception as e:
        st.warning(f"{label_clean}: fit failed ({e})")
        continue

    if model.startswith("4PL"):
        bottom, top, ec50, hill = popt
        fn = lambda xx: four_param_logistic(xx, bottom, top, ec50, hill)
    else:
        top, ec50, hill = popt
        bottom = 0.0
        fn = lambda xx: three_param_logistic(xx, top, ec50, hill)

    # === EC90を計算（EC50 * 9^(1/Hill)）===
    ec90 = calc_ec90(ec50, hill)

    mask = np.isfinite(x) & np.isfinite(y_mean) & (x > 0)
    x_fit, y_fit = x[mask], y_mean[mask]
    y_pred = fn(x_fit)
    r2 = r_squared(y_fit, y_pred)

    xs = np.logspace(np.log10(np.min(x_fit)), np.log10(np.max(x_fit)), 400)
    ys = fn(xs)
    color = colors[i % len(colors)]
    ax.errorbar(x_fit, y_mean[mask], yerr=y_sd[mask], fmt='o', capsize=2, label=label_clean, color=color)
    ax.plot(xs, ys, '-', color=color, label=None)

    if show_ec50_line:
        ax.axvline(ec50, color=color, linestyle='--', linewidth=1)

    # === Fittedデータ ===
    df_fit = pd.DataFrame({"Dataset": label_clean, "X": xs, "Predicted_Y": ys})
    all_fitted.append(df_fit)

    # === Mean/SDデータ ===
    subset_df = data[[x_col] + y_cols].copy()
    subset_df["Dataset"] = label_clean
    subset_df["Mean"] = y_mean
    subset_df["SD"] = y_sd
    subset_df = subset_df[["Dataset", x_col] + y_cols + ["Mean", "SD"]]
    mean_sd_list.append(subset_df)

    # === 結果（EC90付き） ===
    results.append({
        "Dataset": label_clean,
        "Bottom": bottom,
        "Top": top,
        "Hill": hill,
        "EC50": ec50,
        "EC90": ec90,
        "R²": r2
    })

# ==== 軸スケール ====
if all_fitted:
    all_x = np.concatenate([df["X"].values for df in all_fitted])
    all_y = np.concatenate([df["Predicted_Y"].values for df in all_fitted])
    x_min_data = np.nanmin(all_x[all_x > 0]) if np.any(all_x > 0) else 1e-3
    x_max_data = np.nanmax(all_x)
    auto_x_min = max(x_min_data / 10, 1e-6)
    auto_x_max = x_max_data * 10
    auto_y_min = -100000
    auto_y_max = np.nanmax(all_y) * 1.2

    if manual_scale:
        st.sidebar.markdown("#### Manual input")
        x_min = st.sidebar.number_input("X min (log)", value=float(auto_x_min), format="%.6f")
        x_max = st.sidebar.number_input("X max (log)", value=float(auto_x_max), format="%.6f")
        y_min = st.sidebar.number_input("Y min", value=float(auto_y_min), format="%.2f")
        y_max = st.sidebar.number_input("Y max", value=float(auto_y_max), format="%.2f")
    else:
        x_min, x_max, y_min, y_max = auto_x_min, auto_x_max, auto_y_min, auto_y_max

    ax.set_xlim(max(x_min, 1e-6), x_max)
    ax.set_ylim(y_min, y_max)

# ==== 軸フォーマット ====
def sci_formatter(val, pos=None):
    if not np.isfinite(val) or val == 0:
        return "0"
    try:
        exp = int(np.floor(np.log10(abs(val))))
        mant = val / 10**exp
        return f"{mant:g}×10^{exp}"
    except Exception:
        return str(val)

ax.set_xscale("log")
ax.xaxis.set_major_locator(LogLocator(base=10.0, numticks=12))
ax.xaxis.set_minor_formatter(NullFormatter())

if x_axis_format == "scientific (×10^n)":
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, _: sci_formatter(v)))
else:
    ticks = [10**i for i in range(-3, 4)]
    ax.set_xticks(ticks)
    ax.set_xticklabels([f"{t:g}" for t in ticks])

if y_axis_format == "scientific (×10^n)":
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: sci_formatter(v)))
else:
    ax.yaxis.set_major_formatter(ScalarFormatter(useMathText=False))

for spine in ["top", "right"]:
    ax.spines[spine].set_visible(False)
ax.grid(False)
ax.set_xlabel(f"Peptide ({x_unit})")
ax.set_ylabel(response_label)
ax.set_title("Dose–Response")

# ==== 凡例＋EC50リスト ====
legend = ax.legend(frameon=True, loc="upper left", fontsize=10)
legend.get_frame().set_alpha(0.0)
for text in legend.get_texts():
    text.set_fontfamily("Arial")

fig.canvas.draw()
bbox = legend.get_window_extent().transformed(ax.transAxes.inverted())
legend_bottom = bbox.y0

if show_ec50_list and results:
    for j, r in enumerate(results):
        ax.text(0.02, legend_bottom - 0.05 - j*0.05,
                f"{r['Dataset']}: EC50={r['EC50']:.3g} {x_unit},  EC90={r['EC90']:.3g} {x_unit},  R²={r['R²']:.3f}",
                transform=ax.transAxes, fontsize=9, color=colors[j % len(colors)], ha='left')

# ==== 保存 ====
buf_png, buf_pdf = io.BytesIO(), io.BytesIO()
fig.savefig(buf_png, format="png", bbox_inches="tight", dpi=300)
buf_png.seek(0)
fig.savefig(buf_pdf, format="pdf", bbox_inches="tight", dpi=300)
buf_pdf.seek(0)
st.pyplot(fig, clear_figure=True)

# ==== 出力 ====
if results:
    st.subheader("2) Fit Parameters")
    df_results = pd.DataFrame(results)
    st.dataframe(df_results, width=800)

    st.subheader("3) Export")
    st.download_button("Download plot (PDF)", buf_pdf, file_name="dose_response.pdf", mime="application/pdf")

    df_all = pd.concat(all_fitted, ignore_index=True) if all_fitted else None
    df_mean_sd = pd.concat(mean_sd_list, ignore_index=True) if mean_sd_list else None

    st.markdown("---")
    st.subheader("4) Export all results (one Excel file)")
    excel_bytes = to_excel_multi(df_results, df_all, df_mean_sd, buf_png, model_formula)
    st.download_button("Download All Results (Excel, with plot)",
                       excel_bytes,
                       file_name="DoseResponse_AllResults.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if df_mean_sd is not None:
        st.subheader("5) Raw data + Mean/SD Table")
        st.dataframe(df_mean_sd, width=1000)
else:
    st.info("Upload CSVs or use example data to view results.")
